"""
Excel tax report export — country-aware, bilingual (PL/EN).
"""

from __future__ import annotations

import datetime

from src.models import CalculationResult
from src.i18n import t

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_to_excel(
    result: CalculationResult,
    output_path: str = "raport_podatkowy.xlsx",
    year: int | None = None,
) -> str:
    """Generate an Excel workbook with multiple sheets, country and language aware."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    if year is None:
        year = result.timestamp.year - 1

    try:
        from src.countries import get_country
        country = get_country(result.country_code)
        fmt_str = '#,##0.00 "' + country.currency_symbol + '"'
        form_name = country.tax_form_name
        cgt_rate_display = country.effective_cgt_rate_display
        stock_tax = country.calculate_capital_gains_tax(result.total.stock_profit)
        div_tax = country.calculate_dividend_tax(result.total.dividend_gross, result.total.dividend_tax_foreign)
        int_tax = country.calculate_interest_tax(result.total.interest_gross)
    except Exception:
        fmt_str = '#,##0.00 "zł"'
        form_name = "PIT-38"
        cgt_rate_display = "19%"
        stock_tax = max(0, result.total.stock_profit * 0.19)
        div_tax = result.total.dividend_tax_due
        int_tax = result.total.interest_tax_due

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    money_alignment = Alignment(horizontal="right")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===== Sheet 1: Summary =====
    ws = wb.active
    ws.title = t('pdf.summary')

    ws['A1'] = f"{t('pdf.title')} — {form_name} — {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{t('pdf.generated')}: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"

    headers = [t('pdf.category'), t('pdf.income'), t('pdf.cost'), t('pdf.profit'), t('pdf.tax')]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    tr = result.total
    summary_data = [
        [t('pdf.stocks'), tr.stock_income, tr.stock_cost, tr.stock_profit, stock_tax],
        [t('pdf.dividends'), tr.dividend_gross, tr.dividend_tax_foreign, tr.dividend_gross - tr.dividend_tax_foreign, div_tax],
        [t('pdf.interest'), tr.interest_gross, 0, tr.interest_gross, int_tax],
    ]

    for row_idx, row_data in enumerate(summary_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx > 1:
                cell.number_format = fmt_str
                cell.alignment = money_alignment

    total_tax = stock_tax + div_tax + int_tax
    ws.cell(row=9, column=1, value=t('pdf.total_tax')).font = Font(bold=True, size=12)
    total_cell = ws.cell(row=9, column=5, value=total_tax)
    total_cell.font = Font(bold=True, size=12, color="FF0000")
    total_cell.number_format = fmt_str

    for col_letter, width in [('A', 22), ('B', 18), ('C', 18), ('D', 18), ('E', 18)]:
        ws.column_dimensions[col_letter].width = width

    # ===== Sheet 2: Transactions =====
    ws_tx = wb.create_sheet(t('pdf.transactions'))
    tx_headers = [t('pdf.date'), t('pdf.ticker'), t('pdf.type'), t('pdf.quantity'), t('pdf.value'), 'Broker']
    for col, header in enumerate(tx_headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, tx in enumerate(result.transactions[:500], 2):
        ws_tx.cell(row=row_idx, column=1, value=tx.date.strftime('%Y-%m-%d')).border = thin_border
        ws_tx.cell(row=row_idx, column=2, value=tx.ticker).border = thin_border
        ws_tx.cell(row=row_idx, column=3, value=tx.tx_type).border = thin_border
        ws_tx.cell(row=row_idx, column=4, value=tx.qty).border = thin_border
        cell = ws_tx.cell(row=row_idx, column=5, value=tx.total_pln)
        cell.number_format = fmt_str
        cell.border = thin_border
        ws_tx.cell(row=row_idx, column=6, value=tx.source).border = thin_border

    for col_letter, width in [('A', 12), ('B', 15), ('C', 8), ('D', 10), ('E', 15), ('F', 18)]:
        ws_tx.column_dimensions[col_letter].width = width

    # ===== Sheet 3: Open Positions =====
    ws_pos = wb.create_sheet(t('pdf.open_positions'))
    pos_headers = [t('pdf.ticker'), t('pdf.quantity'), t('pdf.avg_cost'), t('pdf.total_value')]
    for col, header in enumerate(pos_headers, 1):
        cell = ws_pos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, pos in enumerate(result.open_positions, 2):
        ws_pos.cell(row=row_idx, column=1, value=pos.get('ticker', '')).border = thin_border
        ws_pos.cell(row=row_idx, column=2, value=pos.get('qty', 0)).border = thin_border
        avg = pos.get('avg_cost_pln', pos.get('avg_cost', 0))
        c = ws_pos.cell(row=row_idx, column=3, value=avg)
        c.number_format = fmt_str
        c.border = thin_border
        c2 = ws_pos.cell(row=row_idx, column=4, value=pos.get('qty', 0) * avg)
        c2.number_format = fmt_str
        c2.border = thin_border

    wb.save(output_path)
    return output_path
